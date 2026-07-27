#!/usr/bin/env python
"""
Excel document generator for Anglerphish reports.
"""

import datetime
import os
import sys
import json
import re
import requests

# Excel library for creating spreadsheets
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
except ImportError:
    sys.stderr.write("Error: openpyxl library is required. Install with: pip install openpyxl\n")
    sys.exit(1)

# Using shared regular expression from word_generator
_SURROGATE_RE = re.compile(r'[\ud800-\udfff]')

# Optional user-agents parsing - reuse from word_generator
try:
    from user_agents import parse
    USER_AGENTS_AVAILABLE = True
except ImportError:
    USER_AGENTS_AVAILABLE = False
    sys.stderr.write("Warning: user_agents library not available. Install with: pip install user-agents\n")

# Optional GeoIP for IP location - reuse from word_generator
try:
    import geoip2.database
    GEOIP_AVAILABLE = True
    GEOIP_DB_PATHS = [
        'GeoLite2-City.mmdb',
        os.path.join(os.path.dirname(__file__), 'GeoLite2-City.mmdb'),
        os.path.join(os.path.expanduser('~'), 'GeoLite2-City.mmdb'),
        '/usr/share/GeoIP/GeoLite2-City.mmdb',
        '/usr/local/share/GeoIP/GeoLite2-City.mmdb'
    ]
    GEOIP_DB_PATH = next((p for p in GEOIP_DB_PATHS if os.path.isfile(p)), None)
    GEOIP_READER = geoip2.database.Reader(GEOIP_DB_PATH) if GEOIP_DB_PATH else None
except ImportError:
    GEOIP_AVAILABLE = False
    GEOIP_READER = None

# Color definitions for Excel
COLORS = {
    'header_blue': '4472C4',  # Header blue
    'sent_green': '1ABC9C',   # Green for sent emails
    'light_blue': '45D6EF',   # Light blue for reporting
    'amber': 'F9BF3B',        # Amber for opened
    'orange': 'F39C12',       # Orange for clicked
    'red': 'F05B4F',          # Red for data submitted
    'dark_orange': 'E67E22',  # Dark orange for replies
    'light_grey': 'D9D9D9',   # Light grey for alternating rows
    'dark_grey': 'BFBFBF',    # Dark grey for alternating rows
    'white': 'FFFFFF',        # White for text
}

def clean_text(s):
    """Strip invalid UTF-16 surrogates from a string for Excel compatibility."""
    if isinstance(s, str):
        return _SURROGATE_RE.sub('', s)
    return str(s) if s is not None else ''

def sanitize_sheet_name(name):
    """
    Sanitize a string to be used as an Excel worksheet name.
    Excel worksheet names cannot contain: [ ] : * ? / \
    They also must not exceed 31 characters.
    """
    if not name:
        return "Sheet"
    
    # Replace invalid characters
    # Replace [ with ( and ] with )
    name = name.replace('[', '(').replace(']', ')')
    # Remove other invalid characters
    invalid_chars = [':', '*', '?', '/', '\\']
    for char in invalid_chars:
        name = name.replace(char, '')
    
    # Ensure it doesn't exceed 31 characters
    if len(name) > 31:
        name = name[:31]
    
    # Remove leading/trailing spaces
    name = name.strip()
    
    # If empty after sanitization, use default
    if not name:
        return "Sheet"
    
    return name

def is_sms_campaign(campaign):
    """Check if a campaign is SMS-based"""
    return campaign.get('type') == 'sms'

def is_generic_campaign(campaign):
    """Check if a campaign is Generic (landing page only, no email/SMS)"""
    return campaign.get('type') == 'generic'

def is_generic_result(result):
    """Check if a result is for a generic campaign link (no email/phone)"""
    return not result.get('email') and not result.get('phone')

def is_sms_result(result):
    """Check if a result is for an SMS target"""
    if result.get('sms_target'):
        return True
    if result.get('phone') and not result.get('email'):
        return True
    return False

def get_contact_field(result):
    """Get the contact identifier (email or phone) for a result"""
    if is_sms_result(result):
        return result.get('phone', 'Unknown')
    return result.get('email', 'Unknown')

def anonymize_phone(phone):
    """Anonymize a phone number by masking the middle digits"""
    if not phone or phone == 'Unknown':
        return phone
    # Remove all non-digit characters for processing
    digits = ''.join(c for c in phone if c.isdigit())
    if len(digits) < 4:
        return phone  # Too short to anonymize meaningfully
    # Show first 2 and last 2 digits, mask the rest
    return phone[:2] + '*' * (len(digits) - 4) + phone[-2:]

def anonymize_email(email):
    """Anonymize an email address"""
    if not email or email == 'Unknown' or '@' not in email:
        return email
    local, domain = email.split('@', 1)
    if len(local) <= 2:
        return email
    return local[0] + '*' * (len(local) - 2) + local[-1] + '@' + domain

def anonymize_ip(ip):
    """Anonymize an IP address by masking the last octet"""
    if not ip or ip == 'Unknown':
        return ip
    # Simple IP anonymization: mask last octet for IPv4
    if '.' in ip and ip.count('.') == 3:
        return '.'.join(ip.split('.')[:-1]) + '.***'
    return ip

def format_date(date_string):
    """Date formatter for mixed ISO/UTC/custom formats - reused from word_generator."""
    if not date_string or date_string == 'N/A':
        return 'N/A'

    try:
        # Strip Z or timezone if present
        if 'Z' in date_string:
            date_string = date_string.rstrip('Z')

        # Handle microseconds with too many digits
        if 'T' in date_string and '.' in date_string:
            date_part, time_part = date_string.split('T')
            if '+' in time_part or '-' in time_part:
                # Has timezone info, split carefully
                if '+' in time_part:
                    time_main, tz = time_part.split('+')
                    tz = '+' + tz
                else:
                    time_main, tz = time_part.rsplit('-', 1)
                    tz = '-' + tz
            else:
                time_main = time_part
                tz = ''

            if '.' in time_main:
                time_clean, micro = time_main.split('.')
                micro = micro[:6]  # truncate microseconds
                time_main = f"{time_clean}.{micro}"

            date_string = f"{date_part}T{time_main}{tz}"

        # Convert to datetime
        dt = datetime.datetime.fromisoformat(date_string)
        return dt.strftime('%Y-%m-%d %H:%M:%S')

    except Exception:
        return date_string  # fallback for invalid or malformed strings

# User Agent information extraction - reused from word_generator
def extract_user_agent(result):
    b = result.get('browser') if isinstance(result.get('browser'), dict) else None
    if b:
        return b.get('user-agent')
    d = result.get('details') if isinstance(result.get('details'), dict) else None
    if d and isinstance(d.get('browser'), dict):
        return d['browser'].get('user-agent')
    return None

def has_user_interaction(result):
    """Check if user had any interaction with the phishing content"""
    return any([
        result.get('opened'),
        result.get('clicked'),
        result.get('submitted_data'),
        result.get('reported'),
        result.get('replied'),
        result.get('status') in ('Email Opened', 'Clicked Link', 'Submitted Data', 'Email Replied', 'Reported Phish')
    ])

def extract_browser_info(result, check_interaction=True):
    """Extract browser info. Returns N/A if no user interaction and check_interaction is True."""
    if check_interaction and not has_user_interaction(result):
        return "N/A"
    ua = extract_user_agent(result)
    if ua and USER_AGENTS_AVAILABLE:
        try:
            parsed = parse(ua)
            return f"{parsed.browser.family} {parsed.browser.version_string}"
        except:
            return ua
    return "Unknown"

def extract_os_info(result, check_interaction=True):
    """Extract OS info. Returns N/A if no user interaction and check_interaction is True."""
    if check_interaction and not has_user_interaction(result):
        return "N/A"
    ua = extract_user_agent(result)
    if ua and USER_AGENTS_AVAILABLE:
        try:
            parsed = parse(ua)
            return f"{parsed.os.family} {parsed.os.version_string}"
        except:
            return ua
    return "Unknown"

def extract_ip_address(result, check_interaction=True):
    """Extract IP address. Returns N/A if no user interaction and check_interaction is True."""
    if check_interaction and not has_user_interaction(result):
        return "N/A"
    ip = result.get('ip')
    if ip:
        return ip
    d = result.get('details') if isinstance(result.get('details'), dict) else {}
    b = d.get('browser') if isinstance(d.get('browser'), dict) else None
    return b.get('address') if b and b.get('address') else 'Unknown'

def is_private_ip(ip):
    if not ip: return False
    if ip.startswith(('10.','192.168.','127.','169.254.','::1','fc00:','fd')):
        return True
    if ip.startswith('172.'):
        try:
            sec = int(ip.split('.')[1])
            return 16 <= sec <=31
        except: pass
    return False

def get_ip_location(ip, result=None):
    """
    Try GeoLite2 → FreeIPAPI → Gophish lat/lon → Unknown.
    Reused from word_generator.
    """
    if not ip or ip == 'Unknown':
        return 'Unknown'

    # 1) Private ranges
    if is_private_ip(ip):
        return 'Internal Network'

    # 2) GeoLite2 database
    if GEOIP_AVAILABLE and GEOIP_READER:
        try:
            resp = GEOIP_READER.city(ip)
            city = resp.city.name or ""
            country = resp.country.name or ""
            if city or country:
                return f"{city}, {country}"
        except Exception:
            pass

    # 3) FreeIPAPI fallback
    if requests:
        try:
            r = requests.get(f"https://freeipapi.com/api/json/{ip}", timeout=5)
            if r.ok:
                data = r.json()
                city   = data.get('cityName')   or data.get('city')   or ""
                region = data.get('regionName') or data.get('region') or ""
                country= data.get('countryName') or data.get('country') or ""
                loc = ", ".join(p for p in (city, region, country) if p)
                if loc:
                    return loc
        except Exception:
            pass

    # 4) Gophish coordinates fallback
    if result and isinstance(result, dict):
        lat = result.get('latitude')
        lon = result.get('longitude')
        if lat is not None and lon is not None:
            return f"{lat}, {lon}"

    return 'Unknown'

# Helper function to extract payload data - reused from word_generator
def extract_payload_from_dict(data_dict, redact=False):
    """Extract submitted payload as key/value text.

    When redact is True the values are replaced with [REDACTED] but the field
    names are kept, so anonymized reports still show which fields were captured
    (e.g. that a password was submitted) without exposing the captured data."""
    result = ""
    if isinstance(data_dict, dict) and 'payload' in data_dict and isinstance(data_dict['payload'], dict):
        for key, value in data_dict['payload'].items():
            if redact:
                result += f'{key}: "[REDACTED]"\n'
            elif isinstance(value, list) and value:
                result += f"{key}: \"{value[0]}\"\n"
            else:
                result += f"{key}: \"{value}\"\n"
    return result

def _parse_iso(s: str) -> datetime.datetime:
    """Strip a trailing 'Z', truncate any fractional‐seconds to 6 digits,
       and preserve a '+HH:MM' or '-HH:MM' offset for fromisoformat.
       Reused from word_generator."""
    s = s.rstrip('Z')
    # split off the timezone offset if present
    tz = ''
    m = re.search(r'([+-]\d{2}:\d{2})$', s)
    if m:
        tz = m.group(1)
        s  = s[: -len(tz)]
    # truncate fractions to 6 digits
    if '.' in s:
        date_part, frac = s.split('.', 1)
        # keep only digits, then pad/truncate to 6
        frac = ''.join(ch for ch in frac if ch.isdigit())[:6].ljust(6, '0')
        s = f"{date_part}.{frac}"
    return datetime.datetime.fromisoformat(s + tz)

def collect_all_deltas(results):
    """
    Calculate time deltas between email delivery and various actions.
    Reused from word_generator.
    """
    click_deltas = []
    subm_deltas  = []
    reply_deltas = []
    report_deltas= []

    mapping = [
        ('clicked_time',    click_deltas),
        ('submitted_time',  subm_deltas),
        ('replied_time',    reply_deltas),
        ('reported_time',   report_deltas),
    ]

    for r in results:
        dt_str = r.get('send_date')
        if not dt_str:
            continue
        try:
            delivered = _parse_iso(dt_str)
        except Exception:
            continue

        for field, lst in mapping:
            ts = r.get(field)
            if not ts:
                continue
            try:
                t = _parse_iso(ts)
                lst.append((t - delivered).total_seconds() / 60)
            except Exception:
                pass

    # failure = minimum of click/submit/reply per user
    failure_deltas = []
    for r in results:
        times = []
        for fld in ('clicked_time','submitted_time','replied_time'):
            ts = r.get(fld)
            if not ts:
                continue
            try:
                t = _parse_iso(ts)
                d = (t - _parse_iso(r['send_date'])).total_seconds() / 60
                times.append(d)
            except:
                pass
        if times:
            failure_deltas.append(min(times))

    return {
        'Click':        click_deltas,
        'Submit Data':  subm_deltas,
        'Reply':        reply_deltas,
        'Report':       report_deltas,
        'Failure':      failure_deltas,
    }

def create_cell_style(fill_color=None, font_color=None, bold=False, alignment=None, border=None):
    """Create a dict of styles for Excel cells"""
    style = {}
    
    if fill_color:
        style['fill'] = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    
    font_opts = {}
    if font_color:
        font_opts['color'] = font_color
    if bold:
        font_opts['bold'] = True
    if font_opts:
        style['font'] = Font(**font_opts)
    
    if alignment:
        style['alignment'] = alignment
    
    if border:
        style['border'] = border
    
    return style

def apply_cell_style(cell, style_dict):
    """Apply a style dict to a cell"""
    for key, value in style_dict.items():
        setattr(cell, key, value)

def style_header_row(worksheet, row_num, start_col, end_col, fill_color=COLORS['header_blue'], text_color=COLORS['white']):
    """Style a header row with the specified colors"""
    for col in range(start_col, end_col + 1):
        cell = worksheet.cell(row=row_num, column=col)
        style = create_cell_style(
            fill_color=fill_color,
            font_color=text_color,
            bold=True,
            alignment=Alignment(horizontal='center', vertical='center', wrap_text=True)
        )
        apply_cell_style(cell, style)

def style_table(worksheet, start_row, end_row, start_col, end_col, 
                header_fill=COLORS['header_blue'], 
                row_fill=COLORS['light_grey'], 
                alt_row_fill=COLORS['dark_grey'],
                header_font=COLORS['white']):
    """Style a table region in the worksheet with alternating row colors"""
    # Style header row
    style_header_row(worksheet, start_row, start_col, end_col, header_fill, header_font)
    
    # Style data rows with alternating colors
    for row in range(start_row + 1, end_row + 1):
        fill_color = row_fill if (row - start_row) % 2 == 1 else alt_row_fill
        for col in range(start_col, end_col + 1):
            cell = worksheet.cell(row=row, column=col)
            style = create_cell_style(
                fill_color=fill_color,
                alignment=Alignment(vertical='center', wrap_text=True)
            )
            apply_cell_style(cell, style)

def auto_adjust_column_width(worksheet):
    """Automatically adjust column widths to fit content"""
    dims = {}
    for row in worksheet.rows:
        for cell in row:
            if cell.value:
                # Get the cell value's length, minimum width is 10
                dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value)) + 2))
    
    # Set column widths, limiting to reasonable values
    for col, width in dims.items():
        width = min(max(width, 10), 50)  # Min 10, max 50 characters
        worksheet.column_dimensions[col].width = width

def generate_excel_document(data, output_path, gdpr_options=None):
    """
    Generate an Excel (.xlsx) report from GoPhish campaign data.
    
    Args:
        data: Dictionary containing campaign data
        output_path: Path to save the Excel file
        gdpr_options: Dictionary with GDPR compliance options
          
    Returns:
        Boolean indicating success
    """
    # Validate data structure
    if not isinstance(data, dict) or 'campaigns' not in data or not data['campaigns']:
        sys.stderr.write("Error: Invalid or empty campaign data\n")
        return False
    
    # Preprocess: inject browser & IP from timeline events into results
    for campaign in data['campaigns']:
        # Build results map using both email and phone as keys
        results_map = {}
        for r in campaign.get('results', []):
            email_key = r.get('email')
            phone_key = r.get('phone')
            if email_key:
                results_map[email_key] = r
            if phone_key:
                results_map[phone_key] = r
        
        for event in campaign.get('timeline', []):
            contact = event.get('email')  # Field name is 'email' but may contain phone for SMS
            ts = event.get('time')    # ISO string
            if contact not in results_map:
                continue

            result = results_map[contact]
            msg = event.get('message')

            # record the moment the message was sent/delivered (email or SMS)
            if msg in ('Email Sent', 'Email Delivered', 'SMS Sent', 'SMS Delivered'):
                result['send_date'] = ts

            # record every action flag
            if msg == 'Email Opened':
                result['opened'] = True
                result['opened_time'] = ts
            elif msg == 'Clicked Link':
                result['clicked'] = True
                result['clicked_time']   = ts
            elif msg == 'Submitted Data':
                result['submitted_data'] = True
                result['submitted_time'] = ts
            elif msg == 'Email Replied':
                result['replied'] = True
                result['replied_time']   = ts
            elif msg == 'Email Reported':
                result['reported'] = True
                result['reported_time']  = ts

            # now inject browser & IP once
            details = event.get('details')
            if isinstance(details, str):
                try:
                    details = json.loads(details)
                except json.JSONDecodeError:
                    details = {}
            
            result['details'] = details
            if 'payload' in details:
                    result['payload'] = details['payload']

            browser = details.get('browser') if isinstance(details, dict) else None
            if isinstance(browser, dict):
                result['browser'] = browser
                if 'address' in browser:
                    result['ip'] = browser['address']
    
    # Create new Excel workbook
    workbook = openpyxl.Workbook()
    
    # Create worksheets: Summary, Per Campaign Unique Results, Per Campaign Results (cumulative),
    # Campaign Overview, Campaign Results, Per-Campaign Event sheets (created later), Browser Statistics, OS Statistics, IP Statistics
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    overview_sheet = workbook.create_sheet("Campaign Information")
    scenario_sheet = workbook.create_sheet("Per Campaign Unique Results")
    cumulative_sheet = workbook.create_sheet("Per Campaign Results")
    results_sheet = workbook.create_sheet("Campaign Results")
    browser_sheet = workbook.create_sheet("Browser Statistics")
    os_sheet = workbook.create_sheet("OS Statistics")
    ip_sheet = workbook.create_sheet("IP Statistics")
    
    # -----------------------------------------
    # Summary Worksheet
    # -----------------------------------------
    summary_sheet['A1'] = "Anglerphish Campaign Report"
    summary_sheet['A1'].font = Font(size=16, bold=True)
    summary_sheet.merge_cells('A1:G1')
    summary_sheet['A1'].alignment = Alignment(horizontal='center')
    
    summary_sheet['A2'] = f"Generated on: {datetime.datetime.now():%Y-%m-%d %H:%M:%S}"
    summary_sheet['A2'].font = Font(italic=True)
    summary_sheet.merge_cells('A2:G2')
    summary_sheet['A2'].alignment = Alignment(horizontal='center')
    
    # Add privacy notice if anonymization is enabled
    current_row = 3
    if gdpr_options and (gdpr_options.get('anonymize_emails', False) or gdpr_options.get('anonymize_ips', False)):
        summary_sheet['A3'] = "Privacy Notice"
        summary_sheet['A3'].font = Font(bold=True, size=12)
        summary_sheet.merge_cells('A3:G3')
        
        privacy_text = "This report has been generated with personal data obfuscated in accordance with privacy requirements. " + \
                      "Email addresses/phone numbers and IP addresses have been partially masked to protect individual privacy."
        summary_sheet['A4'] = privacy_text
        summary_sheet.merge_cells('A4:G4')
        summary_sheet['A4'].alignment = Alignment(wrap_text=True)
        summary_sheet.row_dimensions[4].height = 40
        
        # Apply a light blue background to the privacy notice
        privacy_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        for cell in summary_sheet['A4:G4'][0]:
            cell.fill = privacy_fill
        
        current_row = 6
    
    # Introduction
    summary_sheet[f'A{current_row}'] = "Summary"
    summary_sheet[f'A{current_row}'].font = Font(bold=True, size=14)
    summary_sheet.merge_cells(f'A{current_row}:G{current_row}')
    current_row += 1
    
    summary_text = "This report provides an analysis of the phishing campaign(s) conducted using the Anglerphish platform, " + \
                  "including performance metrics and key user interactions."
    summary_sheet[f'A{current_row}'] = summary_text
    summary_sheet.merge_cells(f'A{current_row}:G{current_row}')
    summary_sheet[f'A{current_row}'].alignment = Alignment(wrap_text=True)
    summary_sheet.row_dimensions[current_row].height = 40
    current_row += 2
    
    # Flatten results
    all_results = []
    in_progress = False
    for c in data['campaigns']:
        all_results.extend(c.get('results', []))
        if c.get('status') not in ('Complete', 'Completed'):
            in_progress = True
    
    # Build a map of unique users with their highest action across ALL campaigns
    # Action levels: 4=submitted, 3=clicked, 2=opened, 1=sent
    user_highest_action = {}  # email -> highest action level
    user_campaigns_failed = {}  # email -> set of campaign names where user failed
    
    for c in data['campaigns']:
        campaign_name = c.get('name', 'Unnamed')
        for r in c.get('results', []):
            email = r.get('email') or r.get('phone', 'Unknown')
            
            # Determine action level for this result
            if r.get('submitted_data') or r.get('status') == 'Submitted Data':
                action_level = 4
            elif r.get('clicked') or r.get('status') == 'Clicked Link':
                action_level = 3
            elif r.get('opened') or r.get('status') == 'Email Opened':
                action_level = 2
            else:
                action_level = 1  # Email Sent only
            
            # Track highest action per user
            if email not in user_highest_action:
                user_highest_action[email] = action_level
            else:
                user_highest_action[email] = max(user_highest_action[email], action_level)
            
            # Track campaigns where user failed (clicked, submitted, or replied)
            if r.get('clicked') or r.get('replied') or r.get('submitted_data'):
                if email not in user_campaigns_failed:
                    user_campaigns_failed[email] = set()
                user_campaigns_failed[email].add(campaign_name)
    
    # Count unique users per action level (mutually exclusive)
    total_unique_users = len(user_highest_action)
    submitted = sum(1 for level in user_highest_action.values() if level == 4)
    clicked = sum(1 for level in user_highest_action.values() if level == 3)
    opened = sum(1 for level in user_highest_action.values() if level == 2)
    sent_only = sum(1 for level in user_highest_action.values() if level == 1)
    
    # For replied and reported, count unique users who did these actions (across all campaigns)
    users_replied = set()
    users_reported = set()
    for r in all_results:
        email = r.get('email') or r.get('phone', 'Unknown')
        if r.get('replied') or r.get('status') == 'Email Replied':
            users_replied.add(email)
        if r.get('reported') or r.get('status') == 'Reported Phish':
            users_reported.add(email)
    replied = len(users_replied)
    reported = len(users_reported)
    
    # Use total_unique_users for percentage calculations
    total = total_unique_users
    
    def pct(p): return round((p/total)*100) if total else 0
    
    # Overall Campaign Results Table
    summary_sheet[f'A{current_row}'] = "Overall Campaign Results"
    summary_sheet[f'A{current_row}'].font = Font(bold=True, size=12)
    summary_sheet.merge_cells(f'A{current_row}:G{current_row}')
    current_row += 1
    
    # Add explanatory note about deduplication
    explanation_text = "Unique users across all campaigns. Each user counted once by highest-risk action."
    summary_sheet[f'A{current_row}'] = explanation_text
    summary_sheet[f'A{current_row}'].font = Font(italic=True, size=9)
    summary_sheet.merge_cells(f'A{current_row}:G{current_row}')
    current_row += 1
    
    # Headers for stats table
    headers = ["Metric", "Count", "Percentage"]
    for col, header in enumerate(headers, 1):
        cell = summary_sheet.cell(row=current_row, column=col)
        cell.value = header
        apply_cell_style(cell, create_cell_style(
            fill_color=COLORS['header_blue'],
            font_color=COLORS['white'],
            bold=True,
            alignment=Alignment(horizontal='center', vertical='center')
        ))
    
    # Create the stats table
    stats_start_row = current_row
    stats_data = [
        ("Total Recipients", total, ""),
        ("Emails Opened", opened, f"{pct(opened)}%"),
        ("Links Clicked", clicked, f"{pct(clicked)}%"),
        ("Data Submitted", submitted, f"{pct(submitted)}%"),
        ("Email Replies", replied, f"{pct(replied)}%"),
        ("Reported Phishing", reported, f"{pct(reported)}%")
    ]
    
    # Colors for each row
    row_colors = [
        COLORS['header_blue'],  # Total Recipients
        COLORS['amber'],        # Emails Opened
        COLORS['orange'],       # Links Clicked
        COLORS['red'],          # Data Submitted
        COLORS['dark_orange'],  # Email Replies
        COLORS['light_blue']    # Reported Phishing
    ]
    
    for i, (label, value, percentage) in enumerate(stats_data):
        current_row += 1
        row = current_row
        
        # Label cell
        cell = summary_sheet.cell(row=row, column=1)
        cell.value = label
        apply_cell_style(cell, create_cell_style(
            fill_color=row_colors[i],
            font_color=COLORS['white'],
            bold=True
        ))
        
        # Value cell
        cell = summary_sheet.cell(row=row, column=2)
        cell.value = value
        apply_cell_style(cell, create_cell_style(
            fill_color=COLORS['light_grey']
        ))
        
        # Percentage cell
        cell = summary_sheet.cell(row=row, column=3)
        cell.value = percentage
        apply_cell_style(cell, create_cell_style(
            fill_color=COLORS['light_grey']
        ))
    
    stats_end_row = current_row
    current_row += 2
    
    # In progress note
    if in_progress:
        note = "Note: Includes in-progress campaigns; data may change as they complete."
        summary_sheet[f'A{current_row}'] = note
        summary_sheet.merge_cells(f'A{current_row}:G{current_row}')
        summary_sheet[f'A{current_row}'].font = Font(italic=True)
        # Add light yellow background to the note
        for cell in summary_sheet[f'A{current_row}:G{current_row}'][0]:
            cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        current_row += 2
    
    # Key Takeaways section
    summary_sheet[f'A{current_row}'] = "Key Takeaways"
    summary_sheet[f'A{current_row}'].font = Font(bold=True, size=14)
    summary_sheet.merge_cells(f'A{current_row}:G{current_row}')
    current_row += 1
    
    # Phish-Prone Percentage
    unique_failures = len({
        r.get('email')
        for r in all_results
        if r.get('clicked') or r.get('replied') or r.get('submitted_data')
    })
    ppp = pct(unique_failures)
    
    takeaway_bullet = f"• Phish-Prone Percentage: {unique_failures} unique users out of {total} ({ppp}%) failed by engaging in risky behavior (clicked, replied, or submitted data)."
    summary_sheet[f'A{current_row}'] = takeaway_bullet
    summary_sheet.merge_cells(f'A{current_row}:G{current_row}')
    current_row += 1
    
    # Time metrics bullets
    actions = collect_all_deltas(all_results)
    bullets = []
    
    def pct(lst, thresh):
        return round(100 * sum(1 for d in lst if d <= thresh) / len(lst)) if lst else 0
    
    # Near-instant risk
    inst = pct(actions['Failure'], 3)
    if inst:
        bullets.append(f"• {inst}% of failures occurred in under 3 min (near-instant risk)")
    
    # Reporting stats - fix the logic to properly count actual reports
    # Count total users who actually reported (from the reported flag, regardless of timing)
    total_reported = sum(1 for r in all_results if r.get('reported') or r.get('status')=='Reported Phish')
    
    # Calculate never reported percentage based on actual reporting behavior
    never_reported_pct = round(((total - total_reported) / total) * 100) if total else 0
    bullets.append(f"• {never_reported_pct}% of users never reported phishing")

    # For timing-based metrics, only use reports with valid timing data
    reps = actions['Report']  # This contains timing deltas for reports with timing data
    if reps:
        # Only show timing metrics when we have timing data
        never24 = round(100 * sum(1 for d in reps if d > 1440) / len(reps))
        bullets.append(f"• {never24}% of users who reported did so after 24 h")
        
        median = sorted(reps)[len(reps)//2]
        bullets.append(f"• Median time-to-report: {round(median)} min")
        fast = pct(reps, 3)
        if fast:
            bullets.append(f"• {fast}% of users reported phishing within 3 min")
    elif total_reported > 0:
        # We have reports but no timing data
        bullets.append(f"• Timing data not available for {total_reported} reports")
    
    # Add bullets to worksheet
    for bullet in bullets:
        summary_sheet[f'A{current_row}'] = bullet
        summary_sheet.merge_cells(f'A{current_row}:G{current_row}')
        current_row += 1
    
    # Time-to-Action Distribution
    current_row += 1
    summary_sheet[f'A{current_row}'] = "Time-to-Action Distribution"
    summary_sheet[f'A{current_row}'].font = Font(bold=True, size=12)
    summary_sheet.merge_cells(f'A{current_row}:G{current_row}')
    current_row += 1
    
    # Time bins
    bins = [
        (float('-inf'), 3, "< 3 min"),  # Use -inf to capture d=0 and any negative deltas
        (3,    15,    "4–15 min"),
        (15,   60,    "16–60 min"),
        (60,  240,    "1–4 h"),
        (240,1440,    "4–24 h"),
        (1440, float('inf'), "> 24 h"),
    ]
    
    # Create headers for time distribution table
    headers = ["Action"] + [label for _, _, label in bins]
    for col, header in enumerate(headers, 1):
        cell = summary_sheet.cell(row=current_row, column=col)
        cell.value = header
        apply_cell_style(cell, create_cell_style(
            fill_color=COLORS['header_blue'],
            font_color=COLORS['white'],
            bold=True,
            alignment=Alignment(horizontal='center', vertical='center')
        ))
    
    time_dist_start_row = current_row
    
    def pct_bin(lst, low, high):
        return round(100 * sum(1 for d in lst if low < d <= high) / len(lst)) if lst else 0
    
    # Filter out "Failure" for the table
    table_actions = {k:v for k,v in actions.items() if k != 'Failure'}
    
    # Define colors for each action type
    action_colors = {
        'Click': COLORS['orange'],
        'Submit Data': COLORS['red'],
        'Reply': COLORS['dark_orange'],
        'Report': COLORS['light_blue']
    }
    
    for action_name, deltas in table_actions.items():
        current_row += 1
        
        # Action name with colored background
        action_cell = summary_sheet.cell(row=current_row, column=1)
        action_cell.value = action_name
        
        # Apply color to action cell
        action_color = action_colors.get(action_name, COLORS['header_blue'])
        apply_cell_style(action_cell, create_cell_style(
            fill_color=action_color,
            font_color=COLORS['white'],
            bold=True,
            alignment=Alignment(horizontal='center', vertical='center')
        ))
        
        # Percentages for each time bin
        for col, (low, high, _) in enumerate(bins, 2):
            summary_sheet.cell(row=current_row, column=col).value = f"{pct_bin(deltas, low, high)}%"
    
    time_dist_end_row = current_row
    
    # Style the time distribution table (data cells only, not action column)
    for row in range(time_dist_start_row + 1, time_dist_end_row + 1):
        fill_color = COLORS['light_grey'] if (row - time_dist_start_row) % 2 == 1 else COLORS['dark_grey']
        # Skip column 1 (action names) which already has colors
        for col in range(2, len(headers) + 1):
            cell = summary_sheet.cell(row=row, column=col)
            apply_cell_style(cell, create_cell_style(
                fill_color=fill_color,
                alignment=Alignment(horizontal='center', vertical='center')
            ))
    
    # Create per-campaign failure table
    current_row += 2
    summary_sheet[f'A{current_row}'] = "Per Campaign Failures"
    summary_sheet[f'A{current_row}'].font = Font(bold=True, size=12)
    summary_sheet.merge_cells(f'A{current_row}:D{current_row}')
    current_row += 1
    
    # Headers for per-campaign failures table
    headers = ["Campaign", "Failures", "Total", "Failure Rate"]
    for col, header in enumerate(headers, 1):
        cell = summary_sheet.cell(row=current_row, column=col)
        cell.value = header
        apply_cell_style(cell, create_cell_style(
            fill_color=COLORS['header_blue'],
            font_color=COLORS['white'],
            bold=True,
            alignment=Alignment(horizontal='center', vertical='center')
        ))
    
    failures_start_row = current_row
    
    # Populate rows
    for campaign in data['campaigns']:
        name = campaign.get('name', 'N/A')
        results = campaign.get('results', [])
        total = len(results)
        unique_fails = len({
            r.get('email')
            for r in results
            if r.get('clicked') or r.get('replied') or r.get('submitted_data')
        })
        # compute failure rate %
        rate = round((unique_fails/total)*100) if total else 0
        
        current_row += 1
        summary_sheet.cell(row=current_row, column=1).value = clean_text(name)
        summary_sheet.cell(row=current_row, column=2).value = unique_fails
        summary_sheet.cell(row=current_row, column=3).value = total
        summary_sheet.cell(row=current_row, column=4).value = f"{rate}%"
    
    failures_end_row = current_row
    
    # Style the failures table
    style_table(summary_sheet, failures_start_row, failures_end_row, 1, len(headers))
    
    # Auto-adjust column widths
    auto_adjust_column_width(summary_sheet)
    
    # -----------------------------------------
    # Per Campaign Unique Results Worksheet
    # -----------------------------------------
    current_row = 1
    scenario_sheet['A1'] = "Per Campaign Unique Results"
    scenario_sheet['A1'].font = Font(size=16, bold=True)
    scenario_sheet.merge_cells('A1:I1')
    scenario_sheet['A1'].alignment = Alignment(horizontal='center')
    current_row += 2
    
    # Headers for the per Campaign Overall results table
    headers = [
        'Campaign Name', 'Total Targets', 'Data Submitted', 'Clicks', 
        'Email Opened', 'Email Sent', 'Email Replied', 'Reported', 
        'Fail % Percentage', 'Report % Percentage'
    ]
    
    # Create header row with specific colors
    header_colors = [
        COLORS['header_blue'],  # Scenario Name
        COLORS['header_blue'],  # Total Targets
        COLORS['red'],          # Data Submitted
        COLORS['orange'],       # Clicks
        COLORS['amber'],        # Email Opened
        COLORS['header_blue'],  # Email Sent
        COLORS['dark_orange'],  # Email Replied
        COLORS['light_blue'],   # Reported
        COLORS['header_blue'],  # Fail % Percentage
        COLORS['header_blue'],  # Report % Percentage
    ]
    
    for col, (header, color) in enumerate(zip(headers, header_colors), 1):
        cell = scenario_sheet.cell(row=current_row, column=col)
        cell.value = header
        apply_cell_style(cell, create_cell_style(
            fill_color=color,
            font_color=COLORS['white'],
            bold=True,
            alignment=Alignment(horizontal='center', vertical='center', wrap_text=True)
        ))
    
    scenario_start_row = current_row
    
    # Populate rows with per-campaign statistics
    for campaign in data['campaigns']:
        current_row += 1
        
        scenario_name = campaign.get('name', 'Unnamed')
        results = campaign.get('results', [])
        
        # Calculate statistics using mutually exclusive counts (highest action only)
        # Hierarchy: Data Submitted > Clicked Link > Email Opened > Email Sent
        total_targets = len(results)
        
        # Mutually exclusive counts - each user counted only in their highest action
        data_submitted = sum(1 for r in results if r.get('submitted_data') or r.get('status')=='Submitted Data')
        clicks = sum(1 for r in results if (r.get('clicked') or r.get('status')=='Clicked Link') and not (r.get('submitted_data') or r.get('status')=='Submitted Data'))
        email_opened = sum(1 for r in results if (r.get('opened') or r.get('status')=='Email Opened') and not (r.get('clicked') or r.get('status')=='Clicked Link'))
        # Email Sent = users who never opened (ensures total = submitted + clicked + opened + sent)
        email_sent = total_targets - data_submitted - clicks - email_opened
        
        # Replied and reported are separate actions, not part of the main funnel
        email_replied = sum(1 for r in results if r.get('replied') or r.get('status')=='Email Replied')
        reported = sum(1 for r in results if r.get('reported') or r.get('status')=='Reported Phish')
        
        # Calculate failures (clicked, replied, or submitted data)
        # Use email or phone for SMS campaigns
        targets_failed = len({
            r.get('email') or r.get('phone', 'Unknown')
            for r in results
            if r.get('clicked') or r.get('replied') or r.get('submitted_data')
        })
        
        # Calculate percentages
        fail_percentage = round((targets_failed / total_targets) * 100, 2) if total_targets else 0
        report_percentage = round((reported / total_targets) * 100, 2) if total_targets else 0
        
        # Populate cells
        scenario_sheet.cell(row=current_row, column=1).value = clean_text(scenario_name)
        scenario_sheet.cell(row=current_row, column=2).value = total_targets
        scenario_sheet.cell(row=current_row, column=3).value = data_submitted
        scenario_sheet.cell(row=current_row, column=4).value = clicks
        scenario_sheet.cell(row=current_row, column=5).value = email_opened
        scenario_sheet.cell(row=current_row, column=6).value = email_sent
        scenario_sheet.cell(row=current_row, column=7).value = email_replied
        scenario_sheet.cell(row=current_row, column=8).value = reported
        scenario_sheet.cell(row=current_row, column=9).value = f"{fail_percentage}%"
        scenario_sheet.cell(row=current_row, column=10).value = f"{report_percentage}%"
        
        # Center align all cells
        for col in range(1, len(headers) + 1):
            cell = scenario_sheet.cell(row=current_row, column=col)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    scenario_end_row = current_row
    
    # Style the table with alternating row colors
    for row in range(scenario_start_row + 1, scenario_end_row + 1):
        fill_color = COLORS['light_grey'] if (row - scenario_start_row) % 2 == 1 else COLORS['dark_grey']
        for col in range(1, len(headers) + 1):
            cell = scenario_sheet.cell(row=row, column=col)
            cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    
    # Freeze panes at row 3 (headers) and column A
    scenario_sheet.freeze_panes = 'B4'  # Freeze rows 1-3 and column A
    
    # Enable filtering at row 3 (headers)
    scenario_sheet.auto_filter.ref = f"A{scenario_start_row}:J{scenario_end_row}"
    
    # Auto-adjust column widths
    auto_adjust_column_width(scenario_sheet)
    
    # -----------------------------------------
    # Per Campaign Results Worksheet (Cumulative)
    # -----------------------------------------
    current_row = 1
    cumulative_sheet['A1'] = "Per Campaign Results (Cumulative)"
    cumulative_sheet['A1'].font = Font(size=16, bold=True)
    cumulative_sheet.merge_cells('A1:I1')
    cumulative_sheet['A1'].alignment = Alignment(horizontal='center')
    current_row += 2
    
    # Headers for the cumulative results table
    cumulative_headers = [
        'Campaign Name', 'Total Targets', 'Data Submitted', 'Clicked Link', 
        'Email Opened', 'Email Sent', 'Email Replied', 'Reported', 
        'Fail % Percentage', 'Report % Percentage'
    ]
    
    # Create header row with specific colors
    cumulative_header_colors = [
        COLORS['header_blue'],  # Campaign Name
        COLORS['header_blue'],  # Total Targets
        COLORS['red'],          # Data Submitted
        COLORS['orange'],       # Clicked Link
        COLORS['amber'],        # Email Opened
        COLORS['sent_green'],   # Email Sent
        COLORS['dark_orange'],  # Email Replied
        COLORS['light_blue'],   # Reported
        COLORS['header_blue'],  # Fail % Percentage
        COLORS['header_blue'],  # Report % Percentage
    ]
    
    for col, (header, color) in enumerate(zip(cumulative_headers, cumulative_header_colors), 1):
        cell = cumulative_sheet.cell(row=current_row, column=col)
        cell.value = header
        apply_cell_style(cell, create_cell_style(
            fill_color=color,
            font_color=COLORS['white'],
            bold=True,
            alignment=Alignment(horizontal='center', vertical='center', wrap_text=True)
        ))
    
    cumulative_start_row = current_row
    
    # Populate rows with per-campaign cumulative statistics
    for campaign in data['campaigns']:
        current_row += 1
        
        campaign_name = campaign.get('name', 'Unnamed')
        results = campaign.get('results', [])
        
        # Calculate cumulative counts (each action includes users who did further actions)
        # These are NOT mutually exclusive - they're cumulative/hierarchical
        # In cumulative mode: opened includes clicked/submitted, clicked includes submitted
        total_targets = len(results)
        
        # Cumulative counts - each level includes users who did further actions
        total_submitted = sum(1 for r in results if r.get('submitted_data') or r.get('status')=='Submitted Data')
        # Clicked = clicked OR submitted (submitting implies clicking)
        total_clicked = sum(1 for r in results if r.get('clicked') or r.get('status')=='Clicked Link' or r.get('submitted_data') or r.get('status')=='Submitted Data')
        # Opened = opened OR clicked OR submitted (clicking implies opening for cumulative)
        total_opened = sum(1 for r in results if r.get('opened') or r.get('status')=='Email Opened' or r.get('clicked') or r.get('status')=='Clicked Link' or r.get('submitted_data') or r.get('status')=='Submitted Data')
        total_replied = sum(1 for r in results if r.get('replied') or r.get('status')=='Email Replied')
        total_reported = sum(1 for r in results if r.get('reported') or r.get('status')=='Reported Phish')
        
        # Calculate failures (clicked, replied, or submitted data)
        # Use email or phone for SMS campaigns
        targets_failed = len({
            r.get('email') or r.get('phone', 'Unknown')
            for r in results
            if r.get('clicked') or r.get('replied') or r.get('submitted_data')
        })
        
        # Calculate percentages
        fail_percentage = round((targets_failed / total_targets) * 100, 2) if total_targets else 0
        report_percentage = round((total_reported / total_targets) * 100, 2) if total_targets else 0
        
        # Populate cells with cumulative counts
        cumulative_sheet.cell(row=current_row, column=1).value = clean_text(campaign_name)
        cumulative_sheet.cell(row=current_row, column=2).value = total_targets
        cumulative_sheet.cell(row=current_row, column=3).value = total_submitted
        cumulative_sheet.cell(row=current_row, column=4).value = total_clicked
        cumulative_sheet.cell(row=current_row, column=5).value = total_opened
        cumulative_sheet.cell(row=current_row, column=6).value = total_targets  # Email Sent = all recipients (100%)
        cumulative_sheet.cell(row=current_row, column=7).value = total_replied
        cumulative_sheet.cell(row=current_row, column=8).value = total_reported
        cumulative_sheet.cell(row=current_row, column=9).value = f"{fail_percentage}%"
        cumulative_sheet.cell(row=current_row, column=10).value = f"{report_percentage}%"
        
        # Center align all cells
        for col in range(1, len(cumulative_headers) + 1):
            cell = cumulative_sheet.cell(row=current_row, column=col)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    cumulative_end_row = current_row
    
    # Style the table with alternating row colors
    for row in range(cumulative_start_row + 1, cumulative_end_row + 1):
        fill_color = COLORS['light_grey'] if (row - cumulative_start_row) % 2 == 1 else COLORS['dark_grey']
        for col in range(1, len(cumulative_headers) + 1):
            cell = cumulative_sheet.cell(row=row, column=col)
            cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    
    # Freeze panes at row 3 (headers) and column A
    cumulative_sheet.freeze_panes = 'B4'  # Freeze rows 1-3 and column A
    
    # Enable filtering at row 3 (headers)
    cumulative_sheet.auto_filter.ref = f"A{cumulative_start_row}:J{cumulative_end_row}"
    
    # Auto-adjust column widths
    auto_adjust_column_width(cumulative_sheet)
    
    # -----------------------------------------
    # Campaign Overview Worksheet
    # -----------------------------------------
    # helper for per‐campaign percentages
    def pct_local(part, whole):
        return round((part/whole)*100) if whole else 0
        
    current_row = 1
    overview_sheet['A1'] = "Campaign Overview"
    overview_sheet['A1'].font = Font(size=16, bold=True)
    overview_sheet.merge_cells('A1:C1')
    overview_sheet['A1'].alignment = Alignment(horizontal='center')
    current_row += 2
    
    # Headers for the campaign overview table
    # Dynamic headers based on whether Generic campaigns are present
    has_generic = any(is_generic_campaign(c) for c in data['campaigns'])
    
    headers = [
        'Campaign ID', 'Campaign Name', 'Campaign Type', 'Created Date', 'Launch Date', 'Completion Date', 
        'Status', 'Email Subject', 'Envelope Sender', 'Phishing URL', 'URL Parameter', 'Redirect URL', 
        'Data Captured', 'Passwords Stored', 'Total Recipients/Links', 'Emails Opened', 
        'Links Clicked', 'Data Submitted', 'Email Replied', 'Reported Phishing'
    ]
    
    # Create header row
    for col, header in enumerate(headers, 1):
        cell = overview_sheet.cell(row=current_row, column=col)
        cell.value = header
        
        # Choose color based on header
        if header == 'Emails Opened':
            color = COLORS['amber']
        elif header == 'Links Clicked':
            color = COLORS['orange']
        elif header == 'Data Submitted':
            color = COLORS['red']
        elif header == 'Email Replied':
            color = COLORS['dark_orange']
        elif header == 'Reported Phishing':
            color = COLORS['light_blue']
        else:
            color = COLORS['header_blue']
        
        apply_cell_style(cell, create_cell_style(
            fill_color=color,
            font_color=COLORS['white'],
            bold=True,
            alignment=Alignment(horizontal='center', vertical='center')
        ))
    
    overview_start_row = current_row
    
    # Populate rows with campaign data
    for campaign in data['campaigns']:
        current_row += 1
        
        # Get campaign stats
        results = campaign.get('results', [])
        sent = len(results)
        is_generic = is_generic_campaign(campaign)
        is_sms = is_sms_campaign(campaign)
        
        # Determine campaign type display
        if is_generic:
            campaign_type = "Generic (Landing Page Only)"
        elif is_sms:
            campaign_type = "SMS"
        else:
            campaign_type = "Email"
        
        # For Generic campaigns, count events from timeline (not unique users)
        if is_generic:
            timeline = campaign.get('timeline', [])
            clicked_c = sum(1 for e in timeline if e.get('message') == 'Clicked Link')
            subm_c = sum(1 for e in timeline if e.get('message') == 'Submitted Data')
            opened_c = 0  # Generic campaigns don't have email opens
            replied_c = 0  # Generic campaigns don't have email replies
            rep_c = 0  # Generic campaigns don't have reports
        else:
            opened_c = sum(1 for r in results if r.get('opened') or r.get('status')=='Email Opened')
            clicked_c = sum(1 for r in results if r.get('clicked') or r.get('status')=='Clicked Link')
            subm_c = sum(1 for r in results if r.get('submitted_data') or r.get('status')=='Submitted Data')
            replied_c = sum(1 for r in results if r.get('replied') or r.get('status')=='Email Replied')
            rep_c = sum(1 for r in results if r.get('reported') or r.get('status')=='Reported Phish')
        
        # Status display
        raw_status = campaign.get('status', 'N/A')
        if raw_status not in ('Complete', 'Completed'):
            total_recip = campaign.get('total_recipients', sent)
            if total_recip:
                pct = round(sent / total_recip * 100)
                status_display = f"{raw_status} ({pct}% – {sent}/{total_recip} sent)"
            else:
                status_display = raw_status
        else:
            status_display = raw_status
        
        # Populate cells (adjusted for Campaign Type column at position 3)
        overview_sheet.cell(row=current_row, column=1).value = clean_text(str(campaign.get('id', 'N/A')))
        overview_sheet.cell(row=current_row, column=2).value = clean_text(campaign.get('name', 'N/A'))
        overview_sheet.cell(row=current_row, column=3).value = campaign_type
        overview_sheet.cell(row=current_row, column=4).value = format_date(campaign.get('created_date', 'N/A'))
        overview_sheet.cell(row=current_row, column=5).value = format_date(campaign.get('launch_date', 'N/A'))
        overview_sheet.cell(row=current_row, column=6).value = format_date(campaign.get('completed_date', 'N/A'))
        overview_sheet.cell(row=current_row, column=7).value = status_display
        # For Generic campaigns, Email Subject and Envelope Sender are N/A
        overview_sheet.cell(row=current_row, column=8).value = 'N/A' if is_generic else clean_text(campaign.get('template_details', {}).get('subject', 'N/A'))
        overview_sheet.cell(row=current_row, column=9).value = 'N/A' if is_generic else clean_text(campaign.get('template_details', {}).get('envelope_sender', 'N/A'))
        overview_sheet.cell(row=current_row, column=10).value = clean_text(campaign.get('phish_url', 'N/A'))
        overview_sheet.cell(row=current_row, column=11).value = clean_text(campaign.get('urlparam', 'rid'))
        overview_sheet.cell(row=current_row, column=12).value = clean_text(campaign.get('page_details', {}).get('redirect_url', 'N/A'))
        overview_sheet.cell(row=current_row, column=13).value = "Yes" if campaign.get('page_details', {}).get('capture_credentials', 0) else "No"
        overview_sheet.cell(row=current_row, column=14).value = "Yes" if campaign.get('page_details', {}).get('capture_passwords', 0) else "No"
        # For Generic campaigns, show "X links" instead of recipients
        if is_generic:
            overview_sheet.cell(row=current_row, column=15).value = f"{sent} links"
        else:
            overview_sheet.cell(row=current_row, column=15).value = str(sent)
        overview_sheet.cell(row=current_row, column=16).value = 'N/A' if is_generic else f"{opened_c} ({pct_local(opened_c, sent)}%)"
        # For Generic campaigns, show total click events
        if is_generic:
            overview_sheet.cell(row=current_row, column=17).value = f"{clicked_c} clicks"
        else:
            overview_sheet.cell(row=current_row, column=17).value = f"{clicked_c} ({pct_local(clicked_c, sent)}%)"
        if is_generic:
            overview_sheet.cell(row=current_row, column=18).value = f"{subm_c} submissions"
        else:
            overview_sheet.cell(row=current_row, column=18).value = f"{subm_c} ({pct_local(subm_c, sent)}%)"
        overview_sheet.cell(row=current_row, column=19).value = 'N/A' if is_generic else f"{replied_c} ({pct_local(replied_c, sent)}%)"
        overview_sheet.cell(row=current_row, column=20).value = 'N/A' if is_generic else f"{rep_c} ({pct_local(rep_c, sent)}%)"
        
        # Center all cells
        for col in range(1, len(headers) + 1):
            cell = overview_sheet.cell(row=current_row, column=col)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    overview_end_row = current_row
    
    # Style the table with alternating row colors
    for row in range(overview_start_row + 1, overview_end_row + 1):
        fill_color = COLORS['light_grey'] if (row - overview_start_row) % 2 == 1 else COLORS['dark_grey']
        for col in range(1, len(headers) + 1):
            cell = overview_sheet.cell(row=row, column=col)
            cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    
    # Freeze panes at row 3 (headers) and columns A-B
    overview_sheet.freeze_panes = 'C4'  # Freeze rows 1-3 and columns A-B
    
    # Enable filtering at row 3 (headers)
    overview_sheet.auto_filter.ref = f"A{overview_start_row}:R{overview_end_row}"
    
    # Auto-adjust column widths
    auto_adjust_column_width(overview_sheet)
    
    # -----------------------------------------
    # Campaign Results Worksheet
    # -----------------------------------------
    current_row = 1
    results_sheet['A1'] = "Campaign Results"
    results_sheet['A1'].font = Font(size=16, bold=True)
    results_sheet.merge_cells('A1:J1')
    results_sheet['A1'].alignment = Alignment(horizontal='center')
    current_row += 2
    
    # Headers for the campaign results table
    headers = [
        'Campaign ID', 'Campaign Name', 'ID (RID)', 'Contact', 'Status', 'IP', 'Location', 
        'Latitude', 'Longitude', 'Send Date', 'Reported', 'Modified Date', 'First Name', 
        'Last Name', 'Position', 'Custom'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = results_sheet.cell(row=current_row, column=col)
        cell.value = header
        apply_cell_style(cell, create_cell_style(
            fill_color=COLORS['header_blue'],
            font_color=COLORS['white'],
            bold=True,
            alignment=Alignment(horizontal='center', vertical='center')
        ))
    
    results_start_row = current_row
    
    # Collect all results
    all_results_details = []
    
    for campaign in data['campaigns']:
        campaign_id = campaign.get('id', 'N/A')
        campaign_name = campaign.get('name', 'Unnamed')
        is_sms = is_sms_campaign(campaign)
        
        for result in campaign.get('results', []):
            # Extract data from result
            contact = get_contact_field(result)
            status = result.get('status', 'N/A')
            
            # Apply GDPR anonymization if enabled
            if gdpr_options and gdpr_options.get('anonymize_emails'):
                if is_sms_result(result):
                    contact = anonymize_phone(contact)
                else:
                    contact = anonymize_email(contact)
            
            # Adjust status display for SMS campaigns
            if is_sms and status == 'Email Sent':
                status = 'SMS Sent'
            
            # For Email Sent/SMS Sent status, there's no user interaction yet, so IP and location should be N/A
            if status in ('Email Sent', 'SMS Sent'):
                ip = 'N/A'
                location = 'N/A'
            else:
                ip = extract_ip_address(result)
                
                # Apply IP anonymization if enabled
                if gdpr_options and gdpr_options.get('anonymize_ips'):
                    ip = anonymize_ip(ip)
                
                location = get_ip_location(ip, result)
            
            send_date = format_date(result.get('send_date', 'N/A'))
            reported = "Yes" if result.get('reported', False) else "No"
            modified_date = format_date(result.get('modified_date', 'N/A'))
            
            # Extract additional data from payload if available
            first_name = result.get('first_name', 'N/A')
            last_name = result.get('last_name', 'N/A')
            position = result.get('position', 'N/A')
            custom = 'N/A'
            
            if 'payload' in result and isinstance(result['payload'], dict):
                payload = result['payload']
                first_name = payload.get('first_name', first_name)
                last_name = payload.get('last_name', last_name)
                position = payload.get('position', position)
                custom = payload.get('custom', 'N/A')
            
            # Extract coordinates - N/A for Email Sent/SMS Sent, otherwise from result
            if status in ('Email Sent', 'SMS Sent'):
                latitude = 'N/A'
                longitude = 'N/A'
            else:
                latitude = result.get('latitude', 'N/A')
                longitude = result.get('longitude', 'N/A')
            
            # Add to results list
            all_results_details.append({
                'campaign_id': campaign_id,
                'campaign_name': campaign_name,
                'rid': result.get('id', 'N/A'),
                'contact': contact,
                'status': status,
                'ip': ip,
                'location': location,
                'latitude': latitude,
                'longitude': longitude,
                'send_date': send_date,
                'reported': reported,
                'modified_date': modified_date,
                'first_name': first_name,
                'last_name': last_name,
                'position': position,
                'custom': custom
            })
    
    # Add results to the worksheet
    for result_detail in all_results_details:
        current_row += 1
        
        results_sheet.cell(row=current_row, column=1).value = clean_text(str(result_detail['campaign_id']))
        results_sheet.cell(row=current_row, column=2).value = clean_text(result_detail['campaign_name'])
        results_sheet.cell(row=current_row, column=3).value = clean_text(str(result_detail['rid']))
        results_sheet.cell(row=current_row, column=4).value = clean_text(result_detail['contact'])
        
        # Status with color coding
        status_cell = results_sheet.cell(row=current_row, column=5)
        status_cell.value = clean_text(result_detail['status'])
        
        # Color code by status
        status = result_detail['status']
        if status in ('Email Sent', 'SMS Sent'):
            color = COLORS['sent_green']
        elif status == 'Email Opened':
            color = COLORS['amber']
        elif status == 'Clicked Link':
            color = COLORS['orange']
        elif status == 'Submitted Data':
            color = COLORS['red']
        elif status == 'Email Replied':
            color = COLORS['dark_orange']
        elif status == 'Reported Phish':
            color = COLORS['light_blue']
        else:
            color = None
        
        if color:
            status_cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            status_cell.font = Font(color=COLORS['white'], bold=True)
        
        results_sheet.cell(row=current_row, column=6).value = clean_text(result_detail['ip'])
        results_sheet.cell(row=current_row, column=7).value = clean_text(result_detail['location'])
        results_sheet.cell(row=current_row, column=8).value = clean_text(str(result_detail['latitude']) if result_detail['latitude'] != 'N/A' else 'N/A')
        results_sheet.cell(row=current_row, column=9).value = clean_text(str(result_detail['longitude']) if result_detail['longitude'] != 'N/A' else 'N/A')
        results_sheet.cell(row=current_row, column=10).value = result_detail['send_date']
        results_sheet.cell(row=current_row, column=11).value = result_detail['reported']
        results_sheet.cell(row=current_row, column=12).value = result_detail['modified_date']
        results_sheet.cell(row=current_row, column=13).value = clean_text(result_detail['first_name'])
        results_sheet.cell(row=current_row, column=14).value = clean_text(result_detail['last_name'])
        results_sheet.cell(row=current_row, column=15).value = clean_text(result_detail['position'])
        results_sheet.cell(row=current_row, column=16).value = clean_text(result_detail['custom'])
    
    results_end_row = current_row
    
    # Apply alternating row colors but preserve status cell color
    for row in range(results_start_row + 1, results_end_row + 1):
        fill_color = COLORS['light_grey'] if (row - results_start_row) % 2 == 1 else COLORS['dark_grey']
        
        for col in range(1, len(headers) + 1):
            # Skip the status column (5) which already has colors
            if col != 5:
                cell = results_sheet.cell(row=row, column=col)
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            else:
                # For status column, keep the color but center the text
                cell = results_sheet.cell(row=row, column=col)
                if not cell.fill.start_color.index or cell.fill.start_color.index == '00000000':
                    cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Freeze panes at row 3 (headers) and columns A-B
    results_sheet.freeze_panes = 'C4'  # Freeze rows 1-3 and columns A-B
    
    # Enable filtering at row 3 (headers)
    results_sheet.auto_filter.ref = f"A{results_start_row}:N{results_end_row}"
    
    # Auto-adjust column widths
    auto_adjust_column_width(results_sheet)  
    
    # -----------------------------------------
    # Per-Campaign Detailed Events Worksheets
    # -----------------------------------------
    for campaign in data['campaigns']:
        campaign_name = campaign.get('name', 'Unnamed')
        campaign_id = campaign.get('id', 'N/A')
        is_sms = is_sms_campaign(campaign)
        is_generic = is_generic_campaign(campaign)
        
        # Create sheet name (Excel has 31 char limit) and sanitize for Excel compatibility
        sheet_name = sanitize_sheet_name(f"Events-{campaign_name}")
        campaign_events_sheet = workbook.create_sheet(sheet_name)
        
        # Title and headers
        current_row = 1
        campaign_events_sheet[f'A{current_row}'] = f"Detailed Events - {campaign_name}"
        campaign_events_sheet[f'A{current_row}'].font = Font(size=16, bold=True)
        campaign_events_sheet.merge_cells(f'A{current_row}:K{current_row}')
        campaign_events_sheet[f'A{current_row}'].alignment = Alignment(horizontal='center')
        current_row += 2
        
        # Dynamic headers based on campaign type
        if is_generic:
            headers = [
                'RID', 'Link Name', 'IP', 'Latitude', 'Longitude',
                'Location', 'Event Type', 'Time', 'Browser', 'OS', 'Details'
            ]
        elif is_sms:
            headers = [
                'RID', 'Phone Number', 'IP', 'Latitude', 'Longitude',
                'Location', 'Event Type', 'Time', 'Browser', 'OS', 'Details'
            ]
        else:
            headers = [
                'RID', 'Email', 'IP', 'Latitude', 'Longitude',
                'Location', 'Event Type', 'Time', 'Browser', 'OS', 'Details'
            ]
        
        # Apply headers
        for col, header in enumerate(headers, 1):
            cell = campaign_events_sheet.cell(row=current_row, column=col)
            cell.value = header
            apply_cell_style(cell, create_cell_style(
                fill_color=COLORS['header_blue'],
                font_color=COLORS['white'],
                bold=True,
                alignment=Alignment(horizontal='center', vertical='center')
            ))
        
        events_start_row = current_row
        
        # Collect events for THIS campaign only
        campaign_events = []
        
        # For Generic campaigns, process ALL events from timeline directly
        if is_generic:
            # Build a map from RId to result for link name lookup
            result_by_rid = {r.get('id', ''): r for r in campaign.get('results', [])}
            
            # Process ALL timeline events
            for event in campaign.get('timeline', []):
                link_id = event.get('email', '')  # For Generic, email field contains RId
                event_time = event.get('time', '')
                msg = event.get('message', '')
                
                # Find the matching result to get link name
                matching_result = result_by_rid.get(link_id, {})
                link_name = matching_result.get('first_name', 'Unknown Link')
                
                # Parse event details
                details = event.get('details')
                if isinstance(details, str):
                    try:
                        details = json.loads(details)
                    except json.JSONDecodeError:
                        details = {}
                
                # Extract IP, browser, OS from event details
                ip = 'N/A'
                browser_info = 'Unknown'
                os_info = 'Unknown'
                latitude = 'N/A'
                longitude = 'N/A'
                
                if details and isinstance(details, dict):
                    browser = details.get('browser', {})
                    if browser:
                        ip = browser.get('address', 'N/A')
                        ua = browser.get('user-agent', '')
                        if ua and USER_AGENTS_AVAILABLE:
                            try:
                                parsed = parse(ua)
                                browser_info = f"{parsed.browser.family} {parsed.browser.version_string}"
                                os_info = f"{parsed.os.family} {parsed.os.version_string}"
                            except:
                                browser_info = ua[:30] if len(ua) > 30 else ua
                                os_info = 'Unknown'
                
                # Apply IP anonymization if enabled
                if gdpr_options and gdpr_options.get('anonymize_ips') and ip != 'N/A':
                    ip = anonymize_ip(ip)
                
                # Build event detail text
                if msg == 'Clicked Link':
                    detail_text = f'Link clicked from IP: {ip}'
                    if browser_info != 'Unknown':
                        detail_text += f' using {browser_info}'
                    campaign_events.append({
                        'contact': link_name,
                        'result_id': link_id,
                        'event_type': 'Clicked Link',
                        'time': event_time,
                        'ip': ip,
                        'browser': browser_info,
                        'os': os_info,
                        'details': detail_text,
                        'latitude': latitude,
                        'longitude': longitude
                    })
                elif msg == 'Submitted Data':
                    # Extract payload (redact captured values when anonymizing)
                    payload_text = ""
                    if details:
                        payload_text = extract_payload_from_dict(
                            details,
                            redact=bool(gdpr_options and gdpr_options.get('anonymize_emails')))
                    detail_text = f'Data submitted from IP: {ip}'
                    if payload_text:
                        detail_text += f' - Data: {payload_text}'
                    campaign_events.append({
                        'contact': link_name,
                        'result_id': link_id,
                        'event_type': 'Submitted Data',
                        'time': event_time,
                        'ip': ip,
                        'browser': browser_info,
                        'os': os_info,
                        'details': detail_text,
                        'latitude': latitude,
                        'longitude': longitude
                    })
        else:
            # Non-Generic campaigns: process events per result
            for result in campaign.get('results', []):
                contact = get_contact_field(result)
                result_id = result.get('id', 'N/A')
                status = result.get('status', 'N/A')
                
                # Apply GDPR anonymization if enabled
                if gdpr_options and gdpr_options.get('anonymize_emails'):
                    if is_sms_result(result):
                        contact = anonymize_phone(contact)
                    else:
                        contact = anonymize_email(contact)
                
                ip = extract_ip_address(result)
                
                # Apply IP anonymization if enabled
                if gdpr_options and gdpr_options.get('anonymize_ips'):
                    ip = anonymize_ip(ip)
                
                browser = extract_browser_info(result)
                os_info = extract_os_info(result)
                
                # Track which events we've added to avoid duplicates
                events_added = set()
                
                # Determine event type label based on campaign type
                sent_label = 'SMS Sent' if is_sms else 'Email Sent'
                
                # Add sent event
                if 'send_date' in result:
                    campaign_events.append({
                        'contact': contact,
                        'result_id': result_id,
                        'event_type': sent_label,
                        'time': result['send_date'],
                        'ip': 'N/A',
                        'browser': 'N/A',
                        'os': 'N/A',
                        'details': f'{sent_label.split()[0]} successfully delivered',
                        'latitude': 'N/A',  # No user interaction yet, so no coordinates
                        'longitude': 'N/A'  # No user interaction yet, so no coordinates
                    })
                    events_added.add(sent_label)
                
                # Add opened event (only for email campaigns)
                if not is_sms and result.get('opened') and 'opened_time' in result:
                    campaign_events.append({
                        'contact': contact,
                        'result_id': result_id,
                        'event_type': 'Email Opened',
                        'time': result['opened_time'],
                        'ip': ip,
                        'browser': browser,
                        'os': os_info,
                        'details': 'User opened the email',
                        'latitude': result.get('latitude', 'N/A'),
                        'longitude': result.get('longitude', 'N/A')
                    })
                    events_added.add('Email Opened')
                
                # Add clicked event
                if result.get('clicked') and 'clicked_time' in result:
                    campaign_events.append({
                        'contact': contact,
                        'result_id': result_id,
                        'event_type': 'Clicked Link',
                        'time': result['clicked_time'],
                        'ip': ip,
                        'browser': browser,
                        'os': os_info,
                        'details': f'User clicked phishing link - Location: {get_ip_location(ip, result)}',
                        'latitude': result.get('latitude', 'N/A'),
                        'longitude': result.get('longitude', 'N/A')
                    })
                    events_added.add('Clicked Link')
                
                # Add submitted data event
                if result.get('submitted_data') and 'submitted_time' in result:
                    payload = ""
                    redact_payload = bool(gdpr_options and gdpr_options.get('anonymize_emails'))

                    # Try to extract payload data
                    if 'details' in result and isinstance(result['details'], dict):
                        payload = extract_payload_from_dict(result['details'], redact=redact_payload)
                    elif 'browser' in result and isinstance(result['browser'], dict):
                        payload = extract_payload_from_dict(result['browser'], redact=redact_payload)
                    elif 'payload' in result and isinstance(result['payload'], dict):
                        for key, value in result['payload'].items():
                            if redact_payload:
                                payload += f'{key}: "[REDACTED]", '
                            elif isinstance(value, list) and value:
                                payload += f"{key}: \"{value[0]}\", "
                            else:
                                payload += f"{key}: \"{value}\", "
                        payload = payload.rstrip(", ")
                    
                    details = f'User submitted data - Location: {get_ip_location(ip, result)}'
                    if payload:
                        details += f' - Data: {payload}'
                    
                    campaign_events.append({
                        'contact': contact,
                        'result_id': result_id,
                        'event_type': 'Submitted Data',
                        'time': result['submitted_time'],
                        'ip': ip,
                        'browser': browser,
                        'os': os_info,
                        'details': details,
                        'latitude': result.get('latitude', 'N/A'),
                        'longitude': result.get('longitude', 'N/A')
                    })
                    events_added.add('Submitted Data')
                
                # Add replied event (only for email campaigns)
                if not is_sms and result.get('replied') and 'replied_time' in result:
                    campaign_events.append({
                        'contact': contact,
                        'result_id': result_id,
                        'event_type': 'Email Replied',
                        'time': result['replied_time'],
                        'ip': ip,
                        'browser': browser,
                        'os': os_info,
                        'details': 'User replied to the phishing email',
                        'latitude': result.get('latitude', 'N/A'),
                        'longitude': result.get('longitude', 'N/A')
                    })
                    events_added.add('Email Replied')
                
                # Add reported event
                if result.get('reported') and 'reported_time' in result:
                    campaign_events.append({
                        'contact': contact,
                        'result_id': result_id,
                        'event_type': 'Reported Phish',
                        'time': result['reported_time'],
                        'ip': ip,
                        'browser': browser,
                        'os': os_info,
                        'details': 'User reported the phishing attempt',
                        'latitude': result.get('latitude', 'N/A'),
                        'longitude': result.get('longitude', 'N/A')
                    })
                    events_added.add('Reported Phish')
        
        # Sort events by time
        try:
            campaign_events.sort(key=lambda x: x['time'] if x['time'] else "")
        except Exception:
            pass
        
        # Populate rows
        for event in campaign_events:
            current_row += 1
            
            campaign_events_sheet.cell(row=current_row, column=1).value = clean_text(str(event.get('result_id', 'N/A')))
            campaign_events_sheet.cell(row=current_row, column=2).value = clean_text(event['contact'])
            campaign_events_sheet.cell(row=current_row, column=3).value = clean_text(event['ip'])
            
            # Latitude and Longitude
            lat = event.get('latitude', 'N/A')
            lon = event.get('longitude', 'N/A')
            campaign_events_sheet.cell(row=current_row, column=4).value = clean_text(str(lat) if lat != 'N/A' else 'N/A')
            campaign_events_sheet.cell(row=current_row, column=5).value = clean_text(str(lon) if lon != 'N/A' else 'N/A')
            
            # Location
            ip = event['ip']
            location = 'N/A'
            if ip and ip != 'N/A':
                location = get_ip_location(ip)
            campaign_events_sheet.cell(row=current_row, column=6).value = clean_text(location)
            
            # Event type with color coding
            cell = campaign_events_sheet.cell(row=current_row, column=7)
            event_type = event['event_type']
            cell.value = event_type
            
            # Color code by event type
            if event_type in ('Email Sent', 'SMS Sent'):
                color = COLORS['sent_green']
            elif event_type == 'Email Opened':
                color = COLORS['amber']
            elif event_type == 'Clicked Link':
                color = COLORS['orange']
            elif event_type == 'Submitted Data':
                color = COLORS['red']
            elif event_type == 'Email Replied':
                color = COLORS['dark_orange']
            elif event_type == 'Reported Phish':
                color = COLORS['light_blue']
            else:
                color = None
            
            if color:
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                cell.font = Font(color=COLORS['white'], bold=True)
            
            # Time
            campaign_events_sheet.cell(row=current_row, column=8).value = format_date(event['time'])
            
            # Browser & OS
            campaign_events_sheet.cell(row=current_row, column=9).value = clean_text(event['browser'])
            campaign_events_sheet.cell(row=current_row, column=10).value = clean_text(event['os'])
            
            # Details
            campaign_events_sheet.cell(row=current_row, column=11).value = clean_text(event['details'])
        
        events_end_row = current_row
        
        # Style the events table
        style_header_row(campaign_events_sheet, events_start_row, 1, len(headers))
        
        # Apply alternating row colors but preserve event type cell color
        for row in range(events_start_row + 1, events_end_row + 1):
            fill_color = COLORS['light_grey'] if (row - events_start_row) % 2 == 1 else COLORS['dark_grey']
            
            for col in range(1, len(headers) + 1):
                # Skip the event type column (7) which already has colors
                if col != 7:
                    cell = campaign_events_sheet.cell(row=row, column=col)
                    cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                else:
                    # For event type column, keep the color but center the text
                    cell = campaign_events_sheet.cell(row=row, column=col)
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Freeze panes at row 3 (headers) and column A
        campaign_events_sheet.freeze_panes = 'B4'
        
        # Enable filtering at row 3 (headers)
        if events_end_row > events_start_row:
            campaign_events_sheet.auto_filter.ref = f"A{events_start_row}:{get_column_letter(len(headers))}{events_end_row}"
        
        # Auto-adjust column widths
        auto_adjust_column_width(campaign_events_sheet)
    
    # -----------------------------------------
    # Browser Statistics Worksheet
    # -----------------------------------------
    current_row = 1
    browser_sheet['A1'] = "Browser Statistics"
    browser_sheet['A1'].font = Font(size=16, bold=True)
    browser_sheet.merge_cells('A1:C1')
    browser_sheet['A1'].alignment = Alignment(horizontal='center')
    current_row += 2
    
    # Headers for the browser statistics table
    headers = ["Browser", "Count", "Percentage"]
    for col, header in enumerate(headers, 1):
        cell = browser_sheet.cell(row=current_row, column=col)
        cell.value = header
        apply_cell_style(cell, create_cell_style(
            fill_color=COLORS['header_blue'],
            font_color=COLORS['white'],
            bold=True,
            alignment=Alignment(horizontal='center', vertical='center')
        ))
    
    browser_start_row = current_row
    
    # Collect browser statistics - only for users who interacted
    browser_counts = {}
    for r in all_results:
        if not has_user_interaction(r):
            continue  # Skip users who never interacted
        b = extract_browser_info(r, check_interaction=False)  # Already checked
        browser_counts[b] = browser_counts.get(b, 0) + 1
    
    # Calculate total for percentage
    total_browsers = sum(browser_counts.values())
    
    # Add browser statistics to the worksheet
    for browser, count in sorted(browser_counts.items(), key=lambda x: -x[1]):
        current_row += 1
        percentage = round((count / total_browsers) * 100) if total_browsers else 0
        
        browser_sheet.cell(row=current_row, column=1).value = clean_text(browser)
        browser_sheet.cell(row=current_row, column=2).value = count
        browser_sheet.cell(row=current_row, column=3).value = f"{percentage}%"
    
    browser_end_row = current_row
    
    # Style the browser statistics table
    style_table(browser_sheet, browser_start_row, browser_end_row, 1, len(headers))
    
    # Freeze panes at row 3 (headers)
    browser_sheet.freeze_panes = 'A4'  # Freeze rows 1-3
    
    # Enable filtering at row 3 (headers)
    browser_sheet.auto_filter.ref = f"A{browser_start_row}:C{browser_end_row}"
    
    # Auto-adjust column widths
    auto_adjust_column_width(browser_sheet)
    
    # -----------------------------------------
    # OS Statistics Worksheet
    # -----------------------------------------
    current_row = 1
    os_sheet['A1'] = "Operating System Statistics"
    os_sheet['A1'].font = Font(size=16, bold=True)
    os_sheet.merge_cells('A1:C1')
    os_sheet['A1'].alignment = Alignment(horizontal='center')
    current_row += 2
    
    # Headers for the OS statistics table
    headers = ["Operating System", "Count", "Percentage"]
    for col, header in enumerate(headers, 1):
        cell = os_sheet.cell(row=current_row, column=col)
        cell.value = header
        apply_cell_style(cell, create_cell_style(
            fill_color=COLORS['header_blue'],
            font_color=COLORS['white'],
            bold=True,
            alignment=Alignment(horizontal='center', vertical='center')
        ))
    
    os_start_row = current_row
    
    # Collect OS statistics - only for users who interacted
    os_counts = {}
    for r in all_results:
        if not has_user_interaction(r):
            continue  # Skip users who never interacted
        o = extract_os_info(r, check_interaction=False)  # Already checked
        os_counts[o] = os_counts.get(o, 0) + 1
    
    # Calculate total for percentage
    total_os = sum(os_counts.values())
    
    # Add OS statistics to the worksheet
    for os_name, count in sorted(os_counts.items(), key=lambda x: -x[1]):
        current_row += 1
        percentage = round((count / total_os) * 100) if total_os else 0
        
        os_sheet.cell(row=current_row, column=1).value = clean_text(os_name)
        os_sheet.cell(row=current_row, column=2).value = count
        os_sheet.cell(row=current_row, column=3).value = f"{percentage}%"
    
    os_end_row = current_row
    
    # Style the OS statistics table
    style_table(os_sheet, os_start_row, os_end_row, 1, len(headers))
    
    # Freeze panes at row 3 (headers)
    os_sheet.freeze_panes = 'A4'  # Freeze rows 1-3
    
    # Enable filtering at row 3 (headers)
    os_sheet.auto_filter.ref = f"A{os_start_row}:C{os_end_row}"
    
    # Auto-adjust column widths
    auto_adjust_column_width(os_sheet)
    
    # -----------------------------------------
    # IP Statistics Worksheet
    # -----------------------------------------
    current_row = 1
    ip_sheet['A1'] = "IP Address and Location Statistics"
    ip_sheet['A1'].font = Font(size=16, bold=True)
    ip_sheet.merge_cells('A1:D1')
    ip_sheet['A1'].alignment = Alignment(horizontal='center')
    current_row += 2
    
    # Headers for the IP statistics table
    headers = ["IP Address", "Location", "Count", "Percentage"]
    for col, header in enumerate(headers, 1):
        cell = ip_sheet.cell(row=current_row, column=col)
        cell.value = header
        apply_cell_style(cell, create_cell_style(
            fill_color=COLORS['header_blue'],
            font_color=COLORS['white'],
            bold=True,
            alignment=Alignment(horizontal='center', vertical='center')
        ))
    
    ip_start_row = current_row
    
    # Collect IP statistics - only for users who interacted
    ip_counts = {}
    ip_locs = {}
    for r in all_results:
        if not has_user_interaction(r):
            continue  # Skip users who never interacted
        ip = extract_ip_address(r, check_interaction=False)  # Already checked
        ip_counts[ip] = ip_counts.get(ip, 0) + 1
        if ip not in ip_locs:
            ip_locs[ip] = get_ip_location(ip, result=r)
    
    # Calculate total for percentage
    total_ips = sum(ip_counts.values())
    
    # Add IP statistics to the worksheet
    for ip, count in sorted(ip_counts.items(), key=lambda x: -x[1]):
        current_row += 1
        percentage = round((count / total_ips) * 100) if total_ips else 0
        
        ip_sheet.cell(row=current_row, column=1).value = clean_text(ip)
        ip_sheet.cell(row=current_row, column=2).value = clean_text(ip_locs.get(ip, 'Unknown'))
        ip_sheet.cell(row=current_row, column=3).value = count
        ip_sheet.cell(row=current_row, column=4).value = f"{percentage}%"
    
    ip_end_row = current_row
    
    # Style the IP statistics table
    style_table(ip_sheet, ip_start_row, ip_end_row, 1, len(headers))
    
    # Freeze panes at row 3 (headers)
    ip_sheet.freeze_panes = 'A4'  # Freeze rows 1-3
    
    # Enable filtering at row 3 (headers)
    ip_sheet.auto_filter.ref = f"A{ip_start_row}:D{ip_end_row}"
    
    # Auto-adjust column widths
    auto_adjust_column_width(ip_sheet)
    
    # -----------------------------------------
    # Repeat Offenders Worksheet (users who failed in 2+ campaigns)
    # -----------------------------------------
    repeat_offenders = {email: campaigns for email, campaigns in user_campaigns_failed.items() if len(campaigns) > 1}
    
    if repeat_offenders and len(data['campaigns']) > 1:
        repeat_sheet = workbook.create_sheet("Repeat Offenders")
        
        current_row = 1
        repeat_sheet['A1'] = "Repeat Offenders"
        repeat_sheet['A1'].font = Font(size=16, bold=True)
        repeat_sheet.merge_cells('A1:D1')
        repeat_sheet['A1'].alignment = Alignment(horizontal='center')
        current_row += 1
        
        # Add explanatory note
        note_text = "Users who failed (clicked, submitted data, or replied) in more than one campaign. These users may require additional security awareness training."
        repeat_sheet[f'A{current_row}'] = note_text
        repeat_sheet[f'A{current_row}'].font = Font(italic=True, size=9)
        repeat_sheet.merge_cells(f'A{current_row}:D{current_row}')
        repeat_sheet[f'A{current_row}'].alignment = Alignment(wrap_text=True)
        repeat_sheet.row_dimensions[current_row].height = 30
        current_row += 1
        
        # Headers for the repeat offenders table
        headers = ["Email/Phone", "Campaigns Failed", "Campaign Names"]
        for col, header in enumerate(headers, 1):
            cell = repeat_sheet.cell(row=current_row, column=col)
            cell.value = header
            apply_cell_style(cell, create_cell_style(
                fill_color='C55A11',  # Orange/Red for warning
                font_color=COLORS['white'],
                bold=True,
                alignment=Alignment(horizontal='center', vertical='center')
            ))
        
        repeat_start_row = current_row
        
        # Add data rows sorted by number of campaigns failed (descending)
        for email, campaigns in sorted(repeat_offenders.items(), key=lambda x: -len(x[1])):
            current_row += 1
            
            # Apply anonymization if enabled
            display_email = email
            if gdpr_options and gdpr_options.get('anonymize_emails'):
                if '@' in email:
                    display_email = anonymize_email(email)
                else:
                    display_email = anonymize_phone(email)
            
            repeat_sheet.cell(row=current_row, column=1).value = clean_text(display_email)
            repeat_sheet.cell(row=current_row, column=2).value = len(campaigns)
            repeat_sheet.cell(row=current_row, column=3).value = clean_text(', '.join(sorted(campaigns)))
            
            # Center align all cells
            for col in range(1, len(headers) + 1):
                cell = repeat_sheet.cell(row=current_row, column=col)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        repeat_end_row = current_row
        
        # Style the table with alternating row colors
        for row in range(repeat_start_row + 1, repeat_end_row + 1):
            fill_color = COLORS['light_grey'] if (row - repeat_start_row) % 2 == 1 else COLORS['dark_grey']
            for col in range(1, len(headers) + 1):
                cell = repeat_sheet.cell(row=row, column=col)
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        
        # Freeze panes at row 4 (after headers)
        repeat_sheet.freeze_panes = 'A5'
        
        # Enable filtering
        if repeat_end_row > repeat_start_row:
            repeat_sheet.auto_filter.ref = f"A{repeat_start_row}:C{repeat_end_row}"
        
        # Auto-adjust column widths
        auto_adjust_column_width(repeat_sheet)
    
    # Save the workbook
    try:
        workbook.save(output_path)
        sys.stderr.write(f"Excel report saved to {output_path}\n")
        return True
    except Exception as e:
        sys.stderr.write(f"Error saving Excel document: {e}\n")
        return False
